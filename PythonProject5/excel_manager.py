import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import logging
from Levenshtein import ratio
import os


class ExcelManager:
    """增强版Excel文件操作管理器"""

    def __init__(self, file_path):
        self.logger = logging.getLogger(__name__)
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path)
        self.logger.info(f"成功加载Excel文件: {file_path}")

        # 样式配置
        self.red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        self.green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        self.highlight_font = Font(bold=True, color='FF0000')
        self.missing_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    def enhanced_fill_names(self, sheet_name, name_col, start_row, files):
        """
        增强版文件名填充功能
        :param sheet_name: 工作表名称
        :param name_col: 文件名列字母
        :param start_row: 起始行号
        :param files: 文件路径列表
        """
        sheet = self.wb[sheet_name]
        col_idx = openpyxl.utils.column_index_from_string(name_col)

        # 收集搜索到的文件名
        search_filenames = {os.path.basename(f) for f in files}
        self.logger.info(f"搜索到 {len(search_filenames)} 个文件名")

        # 收集表格中现有的文件名
        existing_filenames = set()
        for row in range(start_row, sheet.max_row + 1):
            filename = sheet.cell(row, col_idx).value
            if filename:
                existing_filenames.add(filename)

        # 步骤A: 跳过重复文件
        duplicates = search_filenames & existing_filenames
        if duplicates:
            self.logger.info(f"跳过 {len(duplicates)} 个重复文件: {', '.join(list(duplicates)[:3])}...")

        # 步骤B: 处理表格中存在但搜索中没有的文件
        missing_files = existing_filenames - search_filenames
        if missing_files:
            self.logger.info(f"发现 {len(missing_files)} 个缺失文件")
            self._handle_missing_files(sheet, col_idx, missing_files, start_row)

        # 步骤C: 压缩空白行
        self._compress_blank_rows(sheet, start_row)

        # 步骤D: 填充新增文件
        new_files = search_filenames - existing_filenames
        if new_files:
            self.logger.info(f"添加 {len(new_files)} 个新增文件")
            self._fill_new_files(sheet, col_idx, new_files, start_row)

    def _handle_missing_files(self, sheet, col_idx, missing_files, start_row):
        """处理缺失文件（移动或标记）"""
        # 创建缺失文件工作表（如果不存在）
        if "已删除文件" not in self.wb.sheetnames:
            self.wb.create_sheet("已删除文件")

        target_sheet = self.wb["已删除文件"]

        # 复制表头
        if target_sheet.max_row == 0:
            for col in range(1, sheet.max_column + 1):
                target_sheet.cell(1, col).value = sheet.cell(1, col).value

        # 移动缺失文件行
        rows_to_delete = []
        for row in range(sheet.max_row, start_row - 1, -1):
            filename = sheet.cell(row, col_idx).value
            if filename in missing_files:
                # 复制行到目标表
                new_row = target_sheet.max_row + 1
                for col in range(1, sheet.max_column + 1):
                    target_sheet.cell(new_row, col).value = sheet.cell(row, col).value

                # 标记删除行
                rows_to_delete.append(row)

        # 删除已移动的行
        for row in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row)

        self.logger.info(f"已移动 {len(rows_to_delete)} 个缺失文件到'已删除文件'表")

    def _compress_blank_rows(self, sheet, start_row):
        """压缩空白行（步骤C）"""
        blank_rows = []

        # 查找所有空白行
        for row in range(start_row, sheet.max_row + 1):
            is_blank = True
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row, col).value:
                    is_blank = False
                    break
            if is_blank:
                blank_rows.append(row)

        # 从底部开始删除空白行
        for row in sorted(blank_rows, reverse=True):
            sheet.delete_rows(row)

        self.logger.info(f"已删除 {len(blank_rows)} 个空白行")

    def _fill_new_files(self, sheet, col_idx, new_files, start_row):
        """填充新增文件（步骤D）"""
        # 查找第一个空白行
        first_blank_row = None
        for row in range(start_row, sheet.max_row + 1):
            if not sheet.cell(row, col_idx).value:
                first_blank_row = row
                break

        # 如果没找到空白行，则在末尾添加
        if first_blank_row is None:
            first_blank_row = sheet.max_row + 1

        # 填充新文件
        for i, filename in enumerate(new_files):
            sheet.cell(first_blank_row + i, col_idx).value = filename

        self.logger.info(f"已添加 {len(new_files)} 个新文件到第 {first_blank_row} 行")

    def enhanced_fill_texts(self, sheet_name, name_col, text_col, results):
        """
        增强版文本填充功能
        :param sheet_name: 工作表名称
        :param name_col: 文件名列字母
        :param text_col: 文本列字母
        :param results: 识别结果列表
        """
        sheet = self.wb[sheet_name]
        name_col_idx = openpyxl.utils.column_index_from_string(name_col)
        text_col_idx = openpyxl.utils.column_index_from_string(text_col)

        # 创建文件名到文本的映射
        text_map = {res['file']: res['text'] for res in results}

        # 遍历所有行，匹配文件名并填充文本
        filled_count = 0
        for row in range(1, sheet.max_row + 1):
            filename = sheet.cell(row, name_col_idx).value
            if filename and filename in text_map:
                sheet.cell(row, text_col_idx).value = text_map[filename]
                filled_count += 1

        # 自动调整文本列宽度
        max_length = 0
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row, text_col_idx).value
            if cell_value and len(cell_value) > max_length:
                max_length = len(cell_value)

        if max_length > 0:
            sheet.column_dimensions[text_col].width = min(100, max_length + 5)

        self.logger.info(f"已填充 {filled_count} 个识别文本到列 {text_col}")

    def enhanced_compare_texts(self, sheet_name, base_col, compare_col, similarity_threshold=0.85):
        """
        增强版文本比对功能
        :param sheet_name: 工作表名称
        :param base_col: 基准文本列字母
        :param compare_col: 比对文本列字母
        :param similarity_threshold: 相似度阈值
        """
        sheet = self.wb[sheet_name]
        base_idx = openpyxl.utils.column_index_from_string(base_col)
        comp_idx = openpyxl.utils.column_index_from_string(compare_col)

        differences = 0

        for row in range(1, sheet.max_row + 1):
            text_a = str(sheet.cell(row, base_idx).value or "")
            text_b = str(sheet.cell(row, comp_idx).value or "")

            # 计算相似度
            similarity = ratio(text_a, text_b)
            cell = sheet.cell(row, comp_idx)

            # 清除之前的样式
            cell.fill = PatternFill(fill_type=None)
            cell.font = Font()

            if text_b and similarity < similarity_threshold:
                cell.fill = self.red_fill
                cell.font = self.highlight_font
                differences += 1
            elif text_b:
                cell.fill = self.green_fill

        self.logger.info(f"文本比对完成，发现 {differences} 处差异 (阈值: {similarity_threshold})")
        return differences

    def save(self, output_path=None):
        """保存Excel文件"""
        save_path = output_path or self.file_path
        self.wb.save(save_path)
        self.logger.info(f"Excel文件已保存: {save_path}")

    def close(self):
        """关闭工作簿"""
        self.wb.close()
        self.logger.info("Excel工作簿已关闭")