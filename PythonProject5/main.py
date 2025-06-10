import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, StringVar, IntVar, DoubleVar, simpledialog
import tkinter.scrolledtext as scrolledtext
import logging
from stt_engine import STTEngine
import sys
import openpyxl
from openpyxl.styles import Font
import difflib
import json
import time
import threading
from Levenshtein import ratio as similarity_ratio
import re


class AudioToTextTool:
    """音频转文本GUI工具"""

    def load_presets_from_file(self):
        """加载预设文件（自动处理损坏情况）"""
        self.presets = {}  # 重置内存中的预设

        # 添加调试信息
        self.log(f"尝试加载预设文件: {os.path.abspath(self.preset_file)}")

        if not os.path.exists(self.preset_file):
            self.log(f"预设文件不存在: {os.path.abspath(self.preset_file)}", logging.WARNING)
            return

        try:
            with open(self.preset_file, 'r', encoding='utf-8') as f:
                raw_data = f.read()
                if not raw_data.strip():  # 空文件处理
                    self.log("预设文件为空", logging.WARNING)
                    return

                presets = json.loads(raw_data)

                # 强制转换为字典并过滤无效项
                self.presets = {
                    str(name): config
                    for name, config in presets.items()
                    if isinstance(config, dict)
                }

            self.log(f"成功加载 {len(self.presets)} 个预设", logging.INFO)

        except Exception as e:
            self.log(f"预设文件损坏，已重置: {str(e)}", logging.ERROR)
            # 备份损坏的文件
            backup_name = f"{self.preset_file}.bak_{int(time.time())}"
            shutil.copyfile(self.preset_file, backup_name)
            self.presets = {}

    def __init__(self, root):
        self.root = root
        self.root.title("语音转文本工具 V4.0")
        self.root.geometry("1920x1080")
        self.root.iconbitmap(self.resource_path("icon.ico")) if os.path.exists("icon.ico") else None

        # === 关键修复：确保presets在create_widgets之前初始化 ===
        self.preset_file = "audio_to_text_presets.json"
        self.presets = {}
        self.load_presets_from_file()

        # 配置日志
        self.logger = logging.getLogger()
        self.logger.setLevel(logging.INFO)

        # 初始化变量
        self.folder_path = tk.StringVar()
        self.table_path = tk.StringVar()
        self.status_var = tk.StringVar(value="就绪")
        self.found_files = []
        self.displayed_files = []
        self.file_status = {}
        self.results = []
        self.progress_var = tk.DoubleVar(value=0)
        self.is_processing = False
        self.processing_thread = None

        # Excel设置变量
        self.name_col_var = StringVar(value="A")
        self.text_col_var = StringVar(value="B")
        self.compare_col_var = StringVar(value="C")
        self.start_row_var = IntVar(value=2)
        self.similarity_var = DoubleVar(value=0.8)
        self.sheet_var = StringVar(value="")

        # 模型相关变量
        self.model_var = tk.StringVar()
        self.models = {}
        self.model_languages = {}
        self.stt_engine = None  # 初始化为None，不立即加载

        # ... 其他初始化代码 ...
        self.log_dir = "recognition_logs"
        os.makedirs(self.log_dir, exist_ok=True)

        # 1. 仅扫描模型（不加载）
        self.scan_models_lightweight()

        # 2. 如果有模型，设置默认选择但不加载
        if self.models:
            # 优先选择有效的模型
            valid_models = [name for name, info in self.models.items() if info.get('valid', True)]
            if valid_models:
                self.model_var.set(valid_models[0])
                self.log(f"已设置默认模型: {valid_models[0]} (未加载)")
            else:
                self.log("警告: 未找到任何有效模型", logging.WARNING)

        # 创建界面
        self.create_widgets()

    def scan_models_lightweight(self):
        """轻量级模型扫描，只验证基本文件结构不加载模型"""
        models_dir = "models"
        self.models = {}  # 结构: { "显示名称": { "path": 路径, "engine": 类型, "lang": 语言, "valid": 是否有效 } }
        self.model_languages = {}

        if not os.path.exists(models_dir):
            self.log(f"⚠️ 模型目录不存在: {models_dir}", logging.WARNING)
            return

        MODEL_CONFIG = {
            'CN': {'engine': 'vosk', 'lang': 'zh'},
            'EN': {'engine': 'vosk', 'lang': 'en'},
            'WHISPER': {'engine': 'whisper', 'lang': 'multilingual'},
            'MICROSOFT': {'engine': 'microsoft', 'config_file': True},
            'TENCENT': {'engine': 'tencent', 'config_file': True}
        }

        for lang_dir in os.listdir(models_dir):
            lang_path = os.path.abspath(os.path.join(models_dir, lang_dir))
            lang_key = lang_dir.upper()

            if not os.path.isdir(lang_path) or lang_key not in MODEL_CONFIG:
                continue

            config = MODEL_CONFIG[lang_key]

            # 云服务配置验证
            if config.get('config_file'):
                for config_file in os.listdir(lang_path):
                    if not config_file.lower().endswith('.json'):
                        continue

                    config_path = os.path.join(lang_path, config_file)
                    display_name = f"{lang_dir.title()}/{os.path.splitext(config_file)[0]}"
                    is_valid = False

                    try:
                        # 仅验证配置文件是否存在，不加载内容
                        if os.path.exists(config_path):
                            is_valid = True
                    except Exception as e:
                        self.log(f"❌ 配置文件验证失败 {config_file}: {str(e)}", logging.ERROR)

                    self.models[display_name] = {
                        'path': config_path,
                        'engine': config['engine'],
                        'lang': config.get('lang', 'zh'),
                        'valid': is_valid,
                        'loaded': False
                    }
                    self.model_languages[display_name] = config.get('lang', 'zh')
                continue

            # 本地模型验证
            for model_name in os.listdir(lang_path):
                model_path = os.path.join(lang_path, model_name)
                display_name = f"{lang_dir}/{model_name}"
                is_valid = False

                # 快速验证文件/目录是否存在
                if config['engine'] == 'vosk':
                    is_valid = os.path.isdir(model_path)
                elif config['engine'] == 'whisper':
                    is_valid = os.path.exists(model_path)

                self.models[display_name] = {
                    'path': model_path,
                    'engine': config['engine'],
                    'lang': config.get('lang', 'zh'),
                    'valid': is_valid,
                    'loaded': False
                }
                self.model_languages[display_name] = config.get('lang', 'zh')

        if not self.models:
            self.log("⚠️ 未找到任何模型", logging.WARNING)
        else:
            valid_count = sum(1 for m in self.models.values() if m['valid'])
            self.log(f"模型扫描完成，共找到 {len(self.models)} 个模型 ({valid_count} 个有效)")

    def _initialize_variables(self):
        """初始化所有核心属性"""
        # 日志系统最先初始化
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)

        # 确保至少有一个日志处理器
        if not self.logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            self.logger.addHandler(handler)

        # 初始化其他核心变量
        self.preset_file = "audio_to_text_presets.json"
        self.presets = {}
        self.folder_path = tk.StringVar()
        self.table_path = tk.StringVar()
        self.status_var = tk.StringVar(value="就绪")
        self.found_files = []
        self.displayed_files = []
        self.file_status = {}
        self.results = []
        self.progress_var = tk.DoubleVar(value=0)
        self.is_processing = False
        self.processing_thread = None

        # Excel设置变量
        self.name_col_var = StringVar(value="A")
        self.text_col_var = StringVar(value="B")
        self.compare_col_var = StringVar(value="C")
        self.start_row_var = IntVar(value=2)
        self.similarity_var = DoubleVar(value=0.8)
        self.sheet_var = StringVar(value="")

        # 模型相关变量
        self.model_var = tk.StringVar()
        self.models = {}
        self.model_languages = {}

    def _initialize_engine(self):
        """初始化语音识别引擎"""
        try:
            if self.models:
                first_model = list(self.models.keys())[0]
                self.model_var.set(first_model)

                model_info = self.models[first_model]

                # 统一引擎初始化逻辑
                if model_info['engine'] in ('microsoft', 'tencent'):
                    with open(model_info['path'], 'r', encoding='utf-8') as f:
                        config = json.load(f)

                    self.stt_engine = STTEngine(
                        engine_type=model_info['engine'],
                        config=config,
                        lang=self.model_languages[first_model]
                    )
                else:
                    self.stt_engine = STTEngine(
                        model_config=model_info['path'],
                        lang=self.model_languages[first_model],
                        engine_type=model_info['engine']
                    )

                self.log(f"语音识别引擎初始化成功: {first_model}")
            else:
                self.log("警告: 未找到任何模型，请检查models文件夹")
                self.stt_engine = STTEngine()  # 初始化空引擎
        except Exception as e:
            self.log(f"引擎初始化失败: {str(e)}", logging.ERROR)
            messagebox.showerror("引擎错误", str(e))
            self.root.destroy()

    def on_excel_path_changed(self):
        """当Excel路径变更时自动加载工作表"""
        excel_path = self.excel_entry.get()
        if excel_path and os.path.exists(excel_path):
            sheets = self.get_excel_sheets(excel_path)
            self.sheet_combo['values'] = sheets
            if sheets and not self.sheet_var.get():
                self.sheet_var.set(sheets[0])


    def resource_path(self, relative_path):
        """获取资源的绝对路径"""
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def scan_models(self):
        """仅扫描模型目录，验证文件完整性，但不加载模型"""
        models_dir = "models"
        self.models = {}  # 结构: { "显示名称": { "path": 路径, "engine": 类型, "lang": 语言, "valid": 是否有效 } }
        self.model_languages = {}

        if not os.path.exists(models_dir):
            self.log(f"⚠️ 模型目录不存在: {models_dir}", logging.WARNING)
            return

        MODEL_CONFIG = {
            'CN': {'engine': 'vosk', 'lang': 'zh'},
            'EN': {'engine': 'vosk', 'lang': 'en'},
            'WHISPER': {'engine': 'whisper', 'lang': 'multilingual'},
            'MICROSOFT': {'engine': 'microsoft', 'config_file': True},
            'TENCENT': {'engine': 'tencent', 'config_file': True}
        }

        for lang_dir in os.listdir(models_dir):
            lang_path = os.path.abspath(os.path.join(models_dir, lang_dir))
            lang_key = lang_dir.upper()

            if not os.path.isdir(lang_path) or lang_key not in MODEL_CONFIG:
                continue

            config = MODEL_CONFIG[lang_key]

            # 云服务配置验证
            if config.get('config_file'):
                for config_file in os.listdir(lang_path):
                    if not config_file.lower().endswith('.json'):
                        continue

                    config_path = os.path.join(lang_path, config_file)
                    display_name = f"{lang_dir.title()}/{os.path.splitext(config_file)[0]}"
                    is_valid = False

                    try:
                        # 仅验证配置文件完整性，不加载
                        with open(config_path, 'r', encoding='utf-8') as f:
                            config_data = json.load(f)

                        required_fields = {
                            'microsoft': ['api_key', 'region'],
                            'tencent': ['secret_id', 'secret_key']
                        }[config['engine']]

                        is_valid = all(k in config_data for k in required_fields)
                    except Exception as e:
                        self.log(f"❌ 配置文件验证失败 {config_file}: {str(e)}", logging.ERROR)
                        is_valid = False

                    self.models[display_name] = {
                        'path': config_path,
                        'engine': config['engine'],
                        'lang': config_data.get('language', 'zh') if is_valid else 'zh',
                        'valid': is_valid,
                        'loaded': False  # 标记是否已加载
                    }
                    self.model_languages[display_name] = config_data.get('language', 'zh') if is_valid else 'zh'
                continue

            # 本地模型验证
            for model_name in os.listdir(lang_path):
                model_path = os.path.join(lang_path, model_name)
                display_name = f"{lang_dir}/{model_name}"
                is_valid = False

                # Vosk模型验证
                if config['engine'] == 'vosk':
                    is_valid = os.path.isdir(model_path) and any(
                        f.endswith('.mfcc') or f.endswith('.dic')
                        for f in os.listdir(model_path)
                    )

                # Whisper模型验证
                elif config['engine'] == 'whisper':
                    is_valid = (
                            os.path.isdir(model_path) or
                            (os.path.isfile(model_path) and model_name.endswith('.pt'))
                    )

                self.models[display_name] = {
                    'path': model_path,
                    'engine': config['engine'],
                    'lang': config['lang'],
                    'valid': is_valid,
                    'loaded': False  # 标记是否已加载
                }
                self.model_languages[display_name] = config['lang']

        if not self.models:
            self.log("⚠️ 未找到任何有效模型", logging.WARNING)
        else:
            valid_count = sum(1 for m in self.models.values() if m['valid'])
            self.log(f"模型扫描完成，共找到 {len(self.models)} 个模型 ({valid_count} 个有效)")

    def create_widgets(self):
        # 主框架
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 文件选择部分
        file_frame = ttk.LabelFrame(self.main_frame, text="文件选择")
        file_frame.pack(fill=tk.X, padx=5, pady=5)

        # 音频文件夹选择
        ttk.Label(file_frame, text="音频文件夹:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.folder_entry = ttk.Entry(file_frame, width=60)
        self.folder_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        browse_folder_btn = ttk.Button(file_frame, text="浏览", command=self.browse_folder)
        browse_folder_btn.grid(row=0, column=2, padx=5, pady=5)

        # 搜索文件按钮
        search_btn = ttk.Button(file_frame, text="搜索音频文件", command=self.search_audio_files)
        search_btn.grid(row=0, column=3, padx=5, pady=5)

        # 文件列表框
        list_frame = ttk.Frame(file_frame)
        list_frame.grid(row=1, column=0, columnspan=4, padx=5, pady=5, sticky=tk.EW)

        # 滚动条
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 文件列表
        self.file_listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar.set,
            selectmode=tk.EXTENDED,
            height=8,
            width=100
        )
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.file_listbox.yview)

        # Excel文件选择
        ttk.Label(file_frame, text="Excel表格:").grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
        self.excel_entry = ttk.Entry(file_frame, width=60)
        self.excel_entry.grid(row=2, column=1, padx=5, pady=5, sticky=tk.EW)
        browse_excel_btn = ttk.Button(file_frame, text="浏览", command=self.browse_excel_file)
        browse_excel_btn.grid(row=2, column=2, padx=5, pady=5)
        self.excel_entry.bind("<FocusOut>", lambda e: self.on_excel_path_changed())

        # 工作表选择
        ttk.Label(file_frame, text="工作表:").grid(row=3, column=0, padx=5, pady=5, sticky=tk.W)
        self.sheet_combo = ttk.Combobox(file_frame, textvariable=self.sheet_var, width=20)
        self.sheet_combo.grid(row=3, column=1, padx=5, pady=5, sticky=tk.W)
        ttk.Label(file_frame, text="(选择Excel文件后自动加载)").grid(row=3, column=2, padx=5, pady=5, sticky=tk.W)

        # 添加列权重使搜索框可扩展
        file_frame.columnconfigure(1, weight=1)

        # 设置部分 - 使用单独的框架
        settings_frame = ttk.LabelFrame(self.main_frame, text="识别设置")
        settings_frame.pack(fill=tk.X, padx=5, pady=5)

        # 模型选择
        model_frame = ttk.Frame(settings_frame)
        model_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(model_frame, text="选择模型:").pack(side=tk.LEFT, padx=5)

        # 模型下拉菜单
        self.model_combo = ttk.Combobox(
            model_frame,
            textvariable=self.model_var,
            width=30
        )
        self.model_combo.pack(side=tk.LEFT, padx=5)

        # 设置模型下拉菜单的值
        if self.models:
            self.model_combo['values'] = list(self.models.keys())
            self.model_var.set(list(self.models.keys())[0])
        else:
            self.model_combo['values'] = ["未找到模型"]
            self.model_var.set("未找到模型")

        # 添加确认按钮
        self.confirm_btn = ttk.Button(model_frame, text="确认模型",
                                      command=self.confirm_model)
        self.confirm_btn.pack(side=tk.LEFT, padx=5)

        # 状态重置按钮
        self.reset_status_btn = ttk.Button(model_frame, text="重置状态",
                                           command=self.reset_file_status)
        self.reset_status_btn.pack(side=tk.RIGHT, padx=5)

        # Excel设置框架
        excel_settings_frame = ttk.LabelFrame(settings_frame, text="Excel设置")
        excel_settings_frame.pack(fill=tk.X, padx=5, pady=5)

        # 文件名列设置
        ttk.Label(excel_settings_frame, text="文件名列:").grid(row=0, column=0, padx=5, pady=2)
        name_col_entry = ttk.Entry(excel_settings_frame, textvariable=self.name_col_var, width=5)
        name_col_entry.grid(row=0, column=1, padx=5, pady=2)

        # 文本列设置
        ttk.Label(excel_settings_frame, text="文本列:").grid(row=0, column=2, padx=5, pady=2)
        text_col_entry = ttk.Entry(excel_settings_frame, textvariable=self.text_col_var, width=5)
        text_col_entry.grid(row=0, column=3, padx=5, pady=2)

        # 起始行设置
        ttk.Label(excel_settings_frame, text="起始行:").grid(row=0, column=4, padx=5, pady=2)
        start_row_entry = ttk.Entry(excel_settings_frame, textvariable=self.start_row_var, width=5)
        start_row_entry.grid(row=0, column=5, padx=5, pady=2)

        # 对比列设置
        ttk.Label(excel_settings_frame, text="对比列:").grid(row=1, column=0, padx=5, pady=2)
        compare_col_entry = ttk.Entry(excel_settings_frame, textvariable=self.compare_col_var, width=5)
        compare_col_entry.grid(row=1, column=1, padx=5, pady=2)

        # 相似度阈值
        ttk.Label(excel_settings_frame, text="相似度阈值:").grid(row=1, column=2, padx=5, pady=2)
        similarity_entry = ttk.Entry(excel_settings_frame, textvariable=self.similarity_var, width=5)
        similarity_entry.grid(row=1, column=3, padx=5, pady=2)

        # 日志框
        log_frame = ttk.LabelFrame(self.main_frame, text="处理日志")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=15)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_text.config(state=tk.DISABLED)

        # 进度条
        progress_frame = ttk.Frame(self.main_frame)
        progress_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(progress_frame, text="进度:").pack(side=tk.LEFT, padx=5)
        self.progress_bar = ttk.Progressbar(progress_frame,
                                            variable=self.progress_var,
                                            maximum=100,
                                            length=500)
        self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.progress_label = ttk.Label(progress_frame, text="0/0")
        self.progress_label.pack(side=tk.LEFT, padx=5)

        # 按钮区域
        btn_frame = ttk.Frame(self.main_frame)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)

        # 左侧按钮组（操作按钮）
        left_btn_frame = ttk.Frame(btn_frame)
        left_btn_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.start_btn = ttk.Button(left_btn_frame, text="开始识别",
                                    command=self.start_text_generation)
        self.start_btn.pack(side=tk.LEFT, padx=5)

        self.stop_btn = ttk.Button(left_btn_frame, text="停止",
                                   command=self.stop_text_generation,
                                   state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        # 中间按钮组（Excel操作按钮）
        center_btn_frame = ttk.Frame(btn_frame)
        center_btn_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        fill_names_btn = ttk.Button(center_btn_frame, text="填充文件名",
                                    command=self.fill_names)
        fill_names_btn.pack(side=tk.LEFT, padx=5)

        fill_texts_btn = ttk.Button(center_btn_frame, text="填充文本",
                                    command=self.fill_texts)
        fill_texts_btn.pack(side=tk.LEFT, padx=5)

        compare_texts_btn = ttk.Button(center_btn_frame, text="对比文本",
                                       command=self.compare_texts)
        compare_texts_btn.pack(side=tk.LEFT, padx=5)

        export_btn = ttk.Button(center_btn_frame, text="导出报告",
                                command=self.export_report)
        export_btn.pack(side=tk.LEFT, padx=5)

        # 右侧按钮组（工具按钮）
        right_btn_frame = ttk.Frame(btn_frame)
        right_btn_frame.pack(side=tk.RIGHT, fill=tk.X)

        # 修改预设相关按钮
        save_preset_btn = ttk.Button(right_btn_frame, text="创建预设",
                                     command=self.save_current_as_preset)
        save_preset_btn.pack(side=tk.RIGHT, padx=5)

        # 新增保存当前预设按钮
        update_preset_btn = ttk.Button(right_btn_frame, text="保存当前预设",
                                       command=self.save_current_preset)
        update_preset_btn.pack(side=tk.RIGHT, padx=5)

        # 在填充按钮旁边添加新按钮
        fill_from_file_btn = ttk.Button(center_btn_frame, text="通过文件填充文案",
                                        command=self.fill_texts_from_file)
        fill_from_file_btn.pack(side=tk.LEFT, padx=5)


        # 预设选择下拉框
        self.preset_var = tk.StringVar()
        self.preset_combo = ttk.Combobox(right_btn_frame,
                                         textvariable=self.preset_var,
                                         width=15)
        self.preset_combo.pack(side=tk.RIGHT, padx=5)
        self.preset_combo['values'] = list(self.presets.keys())

        load_preset_btn = ttk.Button(right_btn_frame, text="加载预设",
                                     command=self.load_selected_preset)
        load_preset_btn.pack(side=tk.RIGHT, padx=5)

        # 清空日志按钮
        self.clear_btn = ttk.Button(right_btn_frame, text="清空日志",
                                    command=self.clear_log)
        self.clear_btn.pack(side=tk.RIGHT, padx=5)

        # 状态栏
        status_frame = ttk.Frame(self.main_frame)
        status_frame.pack(fill=tk.X, padx=5, pady=2)

        self.status_var = tk.StringVar(value="就绪 | 等待操作")
        status_label = ttk.Label(status_frame, textvariable=self.status_var,
                                 relief=tk.SUNKEN, anchor=tk.W)
        status_label.pack(fill=tk.X)

    def generate_log_entry(self, file_name, status, text):
        """生成标准化日志条目"""
        log_entry = [
            f"文件名: {file_name}",
            f"状态: {status}",
            "识别结果:",
            text,
            "-----"
        ]
        return "\n".join(log_entry)

    def get_excel_sheets(self, excel_path):
        """获取Excel文件的所有工作表名"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            return wb.sheetnames
        except Exception as e:
            self.log(f"读取Excel失败: {str(e)}")
            return []

    def load_selected_preset(self):
        """加载选中预设（带完整性检查）"""
        selected = self.preset_var.get()
        if not selected or selected not in self.presets:
            messagebox.showerror("错误", "请先选择有效预设")
            return

        config = self.presets[selected]

        # 安全设置各个字段
        self.folder_entry.delete(0, tk.END)
        self.folder_entry.insert(0, config.get("folder_path", ""))

        # 新增Excel相关设置 ==============================
        self.excel_entry.delete(0, tk.END)
        excel_path = config.get("excel_path", "")
        self.excel_entry.insert(0, excel_path)

        # ============= 新增的检查 =============
        if excel_path and not os.path.exists(excel_path):
            self.log(f"警告: Excel文件不存在 - {excel_path}")
            messagebox.showwarning("文件不存在", f"Excel文件不存在:\n{excel_path}")
            return  # 可选：是否继续加载其他设置
        # ====================================

        # 如果Excel路径存在，加载工作表
        if excel_path and os.path.exists(excel_path):
            try:
                sheets = self.get_excel_sheets(excel_path)
                self.sheet_combo['values'] = sheets

                # 优先使用预设中保存的工作表名
                saved_sheet = config.get("sheet_name", "")
                if saved_sheet in sheets:
                    self.sheet_var.set(saved_sheet)
                elif sheets:  # 否则默认选第一个
                    self.sheet_var.set(sheets[0])
            except Exception as e:
                self.log(f"加载工作表失败: {str(e)}")
        # ==============================================

        # 设置其他参数
        self.name_col_var.set(config.get("name_col", "A"))
        self.text_col_var.set(config.get("text_col", "B"))
        self.compare_col_var.set(config.get("compare_col", "C"))
        self.start_row_var.set(config.get("start_row", 2))
        self.similarity_var.set(config.get("similarity", 0.8))
        self.model_var.set(config.get("model", ""))

        messagebox.showinfo("成功", f"预设 '{selected}' 已加载")

    def save_presets_to_file(self):
        """保存预设（带数据验证）"""
        try:
            # 调试信息
            self.log(f"尝试保存预设到: {os.path.abspath(self.preset_file)}")
            self.log(f"当前预设内容: {json.dumps(self.presets, indent=2)}")

            with open(self.preset_file, 'w', encoding='utf-8') as f:
                json.dump(
                    self.presets, f,
                    indent=2,
                    ensure_ascii=False
                )
            self.log("预设保存成功")
        except Exception as e:
            self.log(f"保存预设失败: {str(e)}", logging.ERROR)
            messagebox.showerror("保存失败", f"无法保存预设: {str(e)}")

    def save_current_as_preset(self):
        """重命名为：创建预设 - 检查重名并提示覆盖"""
        preset_name = simpledialog.askstring("创建预设", "请输入新预设名称:")
        if not preset_name:
            return

        # 检查是否已存在同名预设
        if preset_name in self.presets:
            if not messagebox.askyesno("确认覆盖", f"预设 '{preset_name}' 已存在，是否覆盖？"):
                return

        # 收集当前设置
        current_settings = {
            "folder_path": self.folder_entry.get(),
            "excel_path": self.excel_entry.get(),
            "sheet_name": self.sheet_var.get(),
            "name_col": self.name_col_var.get(),
            "text_col": self.text_col_var.get(),
            "compare_col": self.compare_col_var.get(),
            "start_row": self.start_row_var.get(),
            "similarity": self.similarity_var.get(),
            "model": self.model_var.get()
        }

        # 保存到预设
        self.presets[preset_name] = current_settings
        self.save_presets_to_file()

        # 更新下拉框
        self.preset_combo['values'] = list(self.presets.keys())
        self.preset_var.set(preset_name)

        messagebox.showinfo("成功", f"预设 '{preset_name}' 已保存")

    def save_current_preset(self):
        """新增：保存当前预设 - 覆盖当前选中的预设"""
        current_preset = self.preset_var.get()
        if not current_preset or current_preset not in self.presets:
            messagebox.showerror("错误", "请先加载要保存的预设")
            return

        # 确认提示
        if not messagebox.askyesno("确认保存", f"确定要覆盖当前预设 '{current_preset}' 吗？"):
            return

        # 收集当前设置
        current_settings = {
            "folder_path": self.folder_entry.get(),
            "excel_path": self.excel_entry.get(),
            "sheet_name": self.sheet_var.get(),
            "name_col": self.name_col_var.get(),
            "text_col": self.text_col_var.get(),
            "compare_col": self.compare_col_var.get(),
            "start_row": self.start_row_var.get(),
            "similarity": self.similarity_var.get(),
            "model": self.model_var.get()
        }

        # 更新预设
        self.presets[current_preset] = current_settings
        self.save_presets_to_file()

        messagebox.showinfo("成功", f"预设 '{current_preset}' 已更新")

    def refresh_preset_combobox(self):
        """强制同步下拉菜单与内存数据"""
        try:
            # 先重新加载预设确保数据最新
            self.load_presets_from_file()

            # 更新下拉菜单
            current_values = list(self.presets.keys())
            self.preset_combo['values'] = current_values

            # 调试信息
            self.log(f"刷新预设下拉菜单，找到 {len(current_values)} 个预设")
            self.log(f"预设列表: {current_values}")

            # 如果没有选中值，设置第一个为默认
            if current_values and not self.preset_var.get():
                self.preset_var.set(current_values[0])
        except Exception as e:
            self.log(f"刷新预设下拉菜单失败: {str(e)}", logging.ERROR)

    def fill_texts_from_file(self):
        """专为当前日志格式优化的填充方法"""
        try:
            # 1. 选择日志文件
            log_file = filedialog.askopenfilename(
                title="选择识别日志文件",
                initialdir=self.log_dir,
                filetypes=[("日志文件", "*.txt"), ("所有文件", "*.*")]
            )
            if not log_file:
                return

            # 2. 解析日志文件
            with open(log_file, 'r', encoding='utf-8') as f:
                log_content = f.read()

            # 使用更精确的正则表达式匹配
            pattern = re.compile(
                r"(\d+)\. 文件名: (.+?\.wav).*?\n"  # 匹配序号和文件名
                r"   状态: 成功\n"
                r"   文本长度: \d+字符\n"
                r"   识别结果:\n([\s\S]+?)\n-{40}",  # 匹配文本内容（含换行）
                re.MULTILINE
            )

            file_text_map = {match.group(2).strip(): match.group(3).strip()
                             for match in pattern.finditer(log_content)}

            if not file_text_map:
                messagebox.showwarning("警告", "未找到符合格式的识别结果")
                return

            # 3. 验证Excel文件
            excel_path = self.excel_entry.get()
            if not excel_path or not os.path.exists(excel_path):
                messagebox.showerror("错误", "请先选择有效的Excel文件")
                return

            # 4. 执行填充
            wb = openpyxl.load_workbook(excel_path)
            sheet = wb[self.sheet_var.get()] if self.sheet_var.get() in wb.sheetnames else wb.active

            name_col = self.name_col_var.get().upper()
            text_col = self.text_col_var.get().upper()
            start_row = self.start_row_var.get()

            filled_count = 0
            for row in range(start_row, sheet.max_row + 1):
                excel_filename = str(sheet[f"{name_col}{row}"].value or "").strip()
                if not excel_filename:
                    continue

                # 精确匹配文件名（带.wav扩展名）
                if excel_filename in file_text_map:
                    sheet[f"{text_col}{row}"] = file_text_map[excel_filename]
                    filled_count += 1
                    # 标记已匹配（避免重复）
                    file_text_map.pop(excel_filename)

            # 5. 保存结果
            try:
                wb.save(excel_path)
                result_msg = f"成功填充 {filled_count} 条文本\n"
                if file_text_map:
                    result_msg += f"未匹配文件: {', '.join(list(file_text_map.keys())[:3])}{'...' if len(file_text_map) > 3 else ''}"
                messagebox.showinfo("完成", result_msg)
            except PermissionError:
                messagebox.showerror("错误", "Excel文件被占用，请关闭后重试")

        except Exception as e:
            messagebox.showerror("错误", f"操作失败: {str(e)}")
            self.log(f"填充错误: {str(e)}", logging.ERROR)

    def fill_texts_intelligently(self, file_text_map):
        """根据文件名匹配，将文本填充到表格对应位置"""
        if not hasattr(self, 'table'):
            messagebox.showwarning("警告", "请先打开Excel文件")
            return

        try:
            # 获取表格数据
            sheet = self.table.sheet
            name_col = self.name_col_var.get().upper()
            text_col = self.text_col_var.get().upper()
            start_row = self.start_row_var.get()

            # 构建文件名到行号的映射
            file_to_row = {}
            for row in range(start_row, sheet.max_row + 1):
                cell_value = str(sheet[f"{name_col}{row}"].value or "").strip()
                if cell_value:
                    # 去掉可能的扩展名进行比较
                    base_name = os.path.splitext(cell_value)[0]
                    file_to_row[base_name] = row

            # 执行填充
            filled_count = 0
            for file_base, text in file_text_map.items():
                if file_base in file_to_row:
                    row = file_to_row[file_base]
                    sheet[f"{text_col}{row}"] = text
                    filled_count += 1

            # 保存并刷新
            self.table.workbook.save(self.table_path.get())
            self.log(f"成功将 {filled_count} 条文本填充到对应位置")

            # 高亮显示已填充的单元格
            if hasattr(self, 'tree'):
                for item in self.tree.get_children():
                    file_name = self.tree.item(item, 'values')[0]
                    base_name = os.path.splitext(str(file_name))[0]
                    if base_name in file_text_map:
                        self.tree.tag_configure('filled', background='lightgreen')
                        self.tree.item(item, tags=('filled',))

        except Exception as e:
            messagebox.showerror("错误", f"填充文本失败: {str(e)}")
            self.log(f"填充文本失败: {str(e)}", logging.ERROR)

    # 文件浏览方法
    def browse_folder(self):
        """浏览并选择文件夹"""
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.folder_entry.delete(0, tk.END)  # 清空输入框
            self.folder_entry.insert(0, folder_path)  # 插入文件夹路径
            self.log(f"已选择文件夹: {folder_path}")

    def browse_excel_file(self):
        """浏览并选择Excel文件并加载工作表"""
        filetypes = [
            ("Excel文件", "*.xlsx *.xls"),
            ("所有文件", "*.*")
        ]
        file_path = filedialog.askopenfilename(filetypes=filetypes)
        if file_path:
            self.excel_entry.delete(0, tk.END)  # 清空输入框
            self.excel_entry.insert(0, file_path)  # 插入文件路径

            # 加载Excel文件的工作表
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheet_names = wb.sheetnames
                self.sheet_combo['values'] = sheet_names
                if sheet_names:
                    # 如果有预设的工作表名称，使用它，否则使用第一个工作表
                    if self.sheet_var.get() and self.sheet_var.get() in sheet_names:
                        self.sheet_var.set(self.sheet_var.get())
                    else:
                        self.sheet_var.set(sheet_names[0])
                    self.log(f"已加载工作表: {', '.join(sheet_names)}")
            except Exception as e:
                self.log(f"加载Excel工作表失败: {str(e)}", logging.ERROR)
                messagebox.showerror("错误", f"无法读取Excel文件: {str(e)}")
                return

            self.log(f"已选择Excel文件: {file_path} | 工作表: {self.sheet_var.get()}")

    def search_audio_files(self):
        """递归搜索文件夹中的音频文件并显示带语言前缀的文件名"""
        folder_path = self.folder_entry.get()
        if not folder_path or not os.path.isdir(folder_path):
            messagebox.showerror("错误", "请先选择有效的文件夹")
            return

        # 支持常见音频格式
        audio_extensions = ['.wav', '.mp3', '.m4a', '.flac', '.ogg']
        self.found_files = []
        self.displayed_files = []

        # 递归搜索所有子文件夹
        for root_dir, _, filenames in os.walk(folder_path):
            for filename in filenames:
                # 检查文件扩展名
                ext = os.path.splitext(filename)[1].lower()
                if ext in audio_extensions:
                    file_path = os.path.join(root_dir, filename)
                    self.found_files.append(file_path)

                    # 获取相对路径（相对于选择的文件夹）
                    rel_path = os.path.relpath(file_path, folder_path)
                    # 提取最高级文件夹作为语言前缀
                    lang_prefix = ""

                    # 如果文件在子文件夹中
                    if os.path.dirname(rel_path):
                        # 取最高级文件夹作为语言前缀
                        lang_prefix = os.path.normpath(rel_path).split(os.sep)[0] + "/"
                    else:
                        # 如果文件在根目录，尝试从文件名提取语言信息
                        if filename.upper().startswith("CN_"):
                            lang_prefix = "CN/"
                        elif filename.upper().startswith("EN_"):
                            lang_prefix = "EN/"
                        else:
                            # 如果没有明确的前缀，使用当前选择的模型语言
                            model_name = self.model_var.get()
                            if model_name in self.model_languages:
                                lang = self.model_languages[model_name]
                                lang_prefix = "CN/" if lang == "zh" else "EN/"

                    # 创建显示名称: 语言前缀 + 文件名
                    display_name = f"{lang_prefix}{filename}"
                    self.displayed_files.append(display_name)

                    # 初始化文件状态
                    if filename not in self.file_status:
                        self.file_status[filename] = False

        # 更新文件列表
        self.file_listbox.delete(0, tk.END)
        for display_name in self.displayed_files:
            filename = display_name.split('/', 1)[-1]  # 去掉语言前缀获取实际文件名
            status = " (已生成)" if self.file_status.get(filename, False) else ""
            self.file_listbox.insert(tk.END, display_name + status)

        self.log(f"找到 {len(self.found_files)} 个音频文件（包含子文件夹）")

    def reset_file_status(self):
        """重置所有文件的生成状态"""
        for filename in self.file_status:
            self.file_status[filename] = False

        # 更新文件列表显示
        self.file_listbox.delete(0, tk.END)
        for display_name in self.displayed_files:
            filename = display_name.split('/', 1)[-1]  # 去掉语言前缀获取实际文件名
            status = " (已生成)" if self.file_status.get(filename, False) else ""
            self.file_listbox.insert(tk.END, display_name + status)

        self.log("已重置所有文件的生成状态")

    # 停止方法
    def stop_text_generation(self):
        """停止文本生成过程"""
        if self.is_processing:
            self.is_processing = False
            self.log("⏹️ 用户请求停止处理")
            self.status_var.set("已停止 | 等待操作")

            # 更新按钮状态
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)

            # 尝试终止处理线程
            if self.processing_thread and self.processing_thread.is_alive():
                try:
                    self.log("等待处理线程安全退出...")
                except Exception as e:
                    self.log(f"停止线程时出错: {str(e)}", logging.ERROR)

    def log(self, message, level=logging.INFO):
        """统一的日志记录方法"""
        if hasattr(self, 'logger'):
            self.logger.log(level, message)

        # 同时写入GUI日志框
        if hasattr(self, 'log_text'):
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, message + "\n")
            self.log_text.see(tk.END)
            self.log_text.config(state=tk.DISABLED)

    def confirm_model(self):
        """确认并加载选定的模型"""
        model_name = self.model_var.get()
        if model_name not in self.models:
            messagebox.showerror("错误", "无效的模型选择")
            return

        model_info = self.models[model_name]

        # 如果已经加载，直接返回
        if model_info.get('loaded', False):
            self.log(f"模型 {model_name} 已加载，无需重复加载")
            return

        # 检查模型有效性
        if not model_info.get('valid', True):
            messagebox.showerror("错误", "所选模型配置无效，请检查模型文件")
            return

        try:
            start_time = time.time()

            # 释放当前已加载的模型（如果有）
            if hasattr(self, 'stt_engine') and self.stt_engine:
                self.stt_engine.release_resources()

            # 加载新模型
            if model_info['engine'] in ('microsoft', 'tencent'):
                with open(model_info['path'], 'r', encoding='utf-8') as f:
                    config = json.load(f)

                self.stt_engine = STTEngine(
                    model_config=config,
                    lang=model_info['lang'],
                    engine_type=model_info['engine']
                )
            else:
                self.stt_engine = STTEngine(
                    model_config=model_info['path'],
                    lang=model_info['lang'],
                    engine_type=model_info['engine']
                )

            # 标记为已加载
            model_info['loaded'] = True
            model_info['valid'] = True  # 加载成功即标记为有效

            load_time = time.time() - start_time
            self.log(f"✅ 模型加载成功: {model_name} (耗时: {load_time:.2f}s)")

            # 测试引擎
            test_result = self.stt_engine.test_model()
            self.log(f"测试结果: {test_result[:50]}...")

        except Exception as e:
            error_msg = f"模型加载失败: {str(e)}"
            self.log(error_msg, logging.ERROR)
            messagebox.showerror("引擎错误", error_msg)
            model_info['valid'] = False  # 标记为无效
            self.stt_engine = None

    def _transcribe_with_tencent(self, audio_path):
        """确保使用腾讯云API进行转录"""
        if not hasattr(self, 'tencent_client') or self.tencent_client is None:
            self.log("腾讯云客户端未初始化", logging.ERROR)
            return ""

        try:
            # 转换音频格式为腾讯云要求的格式
            converted_path = self._convert_audio_for_tencent(audio_path)

            # 调用腾讯云API
            start_time = time.time()
            result = self.tencent_client.transcribe(converted_path)
            cost_time = time.time() - start_time

            self.log(f"腾讯云识别完成 (耗时: {cost_time:.2f}s)")
            return result
        except Exception as e:
            self.log(f"腾讯云识别失败: {str(e)}", logging.ERROR)
            return ""
        finally:
            if 'converted_path' in locals() and os.path.exists(converted_path):
                os.remove(converted_path)

    def _convert_audio_for_tencent(self, input_path):
        """转换为腾讯云要求的音频格式"""
        temp_path = os.path.join(tempfile.gettempdir(), f"tencent_temp_{os.path.basename(input_path)}.wav")

        cmd = [
            "ffmpeg", "-i", input_path,
            "-ar", "16000",  # 采样率16kHz
            "-ac", "1",  # 单声道
            "-acodec", "pcm_s16le",  # 16位有符号PCM
            "-y", temp_path
        ]

        subprocess.run(cmd, check=True, capture_output=True)
        return temp_path

    def start_text_generation(self):
        """启动文本生成流程（完整线程安全版本）"""
        try:
            # === 1. 检查处理状态 ===
            if self.is_processing:
                self.log("⚠️ 已有处理正在进行中")
                return

            # === 2. 获取选中的文件 ===
            selected_indices = self.file_listbox.curselection()
            if not selected_indices:
                messagebox.showwarning("警告", "请先在文件列表中选择要处理的文件")
                return

            # 安全获取文件路径列表
            try:
                selected_files = [self.found_files[i] for i in selected_indices]
                total_files = len(selected_files)
                self.log(f"准备处理 {total_files} 个文件...")
            except IndexError as e:
                self.log(f"文件索引错误: {str(e)}", logging.ERROR)
                messagebox.showerror("错误", "文件列表索引不匹配，请重新搜索文件")
                return

            # === 3. 验证模型 ===
            model_name = self.model_var.get()
            if model_name not in self.models:
                messagebox.showerror("错误", "请先选择有效的模型")
                return

            model_info = self.models[model_name]
            engine_type = model_info['engine']
            language = self.model_languages.get(model_name, 'zh')

            # === 4. 初始化引擎 ===
            try:
                if engine_type in ('microsoft', 'tencent'):
                    with open(model_info['path'], 'r', encoding='utf-8') as f:
                        config = json.load(f)
                    self.stt_engine = STTEngine(
                        model_config=model_info['path'],
                        lang=language,
                        engine_type=engine_type,
                        config=config
                    )
                else:
                    self.stt_engine = STTEngine(
                        model_config=model_info['path'],
                        lang=language,
                        engine_type=engine_type
                    )
            except Exception as e:
                error_msg = f"引擎初始化失败: {str(e)}"
                self.log(error_msg, logging.ERROR)
                messagebox.showerror("引擎错误", error_msg)
                return

            # === 5. 准备处理 ===
            self.is_processing = True
            self.results = []  # 清空之前的结果
            self.progress_var.set(0)

            # UI状态更新（必须通过root.after保证线程安全）
            self.root.after(0, lambda: [
                self.start_btn.config(state=tk.DISABLED),
                self.stop_btn.config(state=tk.NORMAL),
                self.status_var.set(f"准备处理 {total_files} 个文件..."),
                self.progress_label.config(text=f"0/{total_files}")
            ])

            # === 6. 启动处理线程 ===
            self.processing_thread = threading.Thread(
                target=self._process_files_thread,
                args=(selected_files,),
                daemon=True
            )
            self.processing_thread.start()

            # === 7. 启动进度监控 ===
            self._start_progress_monitor(total_files)

            # 在处理完成后添加日志记录
            if self.results:
                self._save_generation_log(selected_files)

        except Exception as e:
            self.log(f"启动过程中发生未预期错误: {str(e)}", logging.ERROR)
            messagebox.showerror("系统错误", f"程序初始化失败: {str(e)}")
            self._reset_processing_state()

    def _reset_processing_state(self):
        """重置处理状态"""
        self.is_processing = False
        self.root.after(0, lambda: [
            self.start_btn.config(state=tk.NORMAL),
            self.stop_btn.config(state=tk.DISABLED),
            self.status_var.set("就绪 | 发生错误")
        ])

    def _process_files_thread(self, file_list):
        """实际处理文件的线程方法"""
        try:
            for idx, file_path in enumerate(file_list, 1):
                if not self.is_processing:
                    break

                # 更新进度（线程安全）
                self.root.after(0, self._update_progress, idx, len(file_list), os.path.basename(file_path))

                try:
                    # 执行转录
                    text = self.stt_engine.transcribe(file_path)
                    filename = os.path.basename(file_path)

                    if text:
                        self.results.append({
                            'file': filename,
                            'text': text,
                            'duration': self._get_audio_duration(file_path)
                        })
                        self.file_status[filename] = True
                        self.log(f"✅ [{idx}/{len(file_list)}] {filename} 转录成功")
                    else:
                        self.log(f"⚠️ [{idx}/{len(file_list)}] {filename} 无转录结果", logging.WARNING)

                except Exception as e:
                    self.log(f"❌ [{idx}/{len(file_list)}] 处理失败: {str(e)}", logging.ERROR)

        finally:
            self.is_processing = False
            self.root.after(0, self._finish_processing)

    def _update_progress(self, processed, total, current_file):
        """线程安全的进度更新"""
        progress = (processed / total) * 100
        self.progress_var.set(progress)
        self.progress_label.config(text=f"{processed}/{total}")
        self.status_var.set(f"处理中: {current_file[:30]}... ({progress:.1f}%)")

    def _finish_processing(self):
        """处理完成后的清理工作"""
        success_count = len([r for r in self.results if r['text']])
        total = self.progress_label.cget("text").split("/")[-1]

        # 更新UI状态
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.status_var.set(f"完成: {success_count}/{total} 成功")

        # 刷新文件列表状态
        for i, display_name in enumerate(self.displayed_files):
            filename = display_name.split('/')[-1].split(' (')[0]
            if filename in self.file_status:
                status = " (已生成)" if self.file_status[filename] else ""
                self.file_listbox.delete(i)
                self.file_listbox.insert(i, f"{display_name.split(' (')[0]}{status}")

        # 显示完成提示
        messagebox.showinfo(
            "处理完成",
            f"成功处理 {success_count}/{total} 个文件\n"
            f"结果已保存在内存中，可点击'填充文本'导出到Excel"
        )

    def _update_progress(self, processed, total, current_file):
        """线程安全的进度更新"""
        progress = (processed / total) * 100
        self.progress_var.set(progress)
        self.progress_label.config(text=f"{processed}/{total}")
        self.status_var.set(f"处理中: {current_file[:30]}...")

    def _finish_processing(self):
        """处理完成后的清理工作"""
        success_count = len([r for r in self.results if r['text']])
        total = self.progress_label.cget("text").split("/")[-1]

        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.status_var.set(f"完成: {success_count}/{total} 成功")

        # 刷新文件列表状态
        for i, display_name in enumerate(self.displayed_files):
            filename = display_name.split('/')[-1].split(' (')[0]
            if filename in self.file_status:
                status = " (已生成)" if self.file_status[filename] else ""
                self.file_listbox.delete(i)
                self.file_listbox.insert(i, f"{display_name.split(' (')[0]}{status}")

        messagebox.showinfo(
            "处理完成",
            f"成功处理 {success_count}/{total} 个文件\n"
            f"结果已保存在内存中，可点击'填充文本'导出到Excel"
        )

    def _start_progress_monitor(self, total_files):
        """修正后的进度监控"""

        def monitor():
            while self.is_processing:
                time.sleep(0.1)  # 现在可以正确使用
                processed = len(self.results)
                progress = (processed / total_files) * 100
                self.root.after(0, lambda: [
                    self.progress_var.set(progress),
                    self.status_var.set(f"处理中: {progress:.1f}%")
                ])

        threading.Thread(target=monitor, daemon=True).start()

    def _get_audio_duration(self, file_path):
        """获取音频时长（示例实现）"""
        try:
            # 实际实现需要根据音频库获取时长
            return "N/A"
        except:
            return "N/A"

    def process_audio_files(self, file_list):
            """批量处理音频文件 - 不再生成单独的TXT文件"""

            if not hasattr(self, 'stt_engine') or self.stt_engine is None:
                self.log("❌ 语音引擎未初始化")
                return

            self.results = []  # 重置结果
            total_files = len(file_list)
            processed_count = 0

            for file_path in file_list:
                if not self.is_processing:
                    break

                try:
                    filename = os.path.basename(file_path)
                    self.log(f"🚀 开始处理: {filename}")

                    # 更新状态
                    self.root.after(0, self.update_progress, processed_count, total_files, filename)

                    # 转录音频
                    text = self.stt_engine.transcribe(file_path)

                    if text:
                        self.log(f"✅ 转录结果: {text}")

                        # 记录结果到内存（不再保存为单独的TXT文件）
                        result = {'file': filename, 'text': text, 'duration': "N/A"}
                        self.results.append(result)

                        # 更新文件状态
                        self.file_status[filename] = True

                        # 更新列表显示
                        self.root.after(0, self.update_file_status_in_list, filename)
                    else:
                        self.log("❌ 转录失败，无结果返回")

                    processed_count += 1

                except Exception as e:
                    self.log(f"🔥 处理过程中出错: {str(e)}")
                    processed_count += 1

            # 处理完成
            self.is_processing = False
            self.root.after(0, self.reset_ui_state)
            self.log(f"处理完成! 成功处理 {len(self.results)}/{total_files} 个文件")
            messagebox.showinfo("完成", f"处理完成! 成功处理 {len(self.results)}/{total_files} 个文件")

    def update_file_status_in_list(self, filename):
            """更新列表框中文件的显示状态"""
            # 找到对应的显示名称
            for i, display_name in enumerate(self.displayed_files):
                if filename in display_name:
                    # 更新列表框中的这一行
                    status = " (已生成)" if self.file_status.get(filename, False) else ""
                    self.file_listbox.delete(i)
                    self.file_listbox.insert(i, display_name + status)
                    break

    def update_progress(self, processed, total, filename):
            """更新进度显示"""
            progress = (processed + 1) / total * 100
            self.progress_var.set(progress)
            self.progress_label.config(text=f"{processed + 1}/{total}")
            self.status_var.set(f"处理中: {processed + 1}/{total} - {filename}")

    def reset_ui_state(self):
            """重置UI状态"""
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
            self.status_var.set("就绪 | 处理完成")

    def get_worksheet(self, wb):
            """获取选定的工作表"""
            sheet_name = self.sheet_var.get()
            if sheet_name and sheet_name in wb.sheetnames:
                return wb[sheet_name]
            else:
                # 如果没有指定工作表或工作表不存在，使用第一个工作表
                self.log(f"警告: 工作表 '{sheet_name}' 不存在，使用第一个工作表")
                return wb.active

    def fill_names(self):
            """增强版文件名填充功能 - 使用选定工作表"""
            excel_path = self.excel_entry.get()
            if not excel_path or not os.path.exists(excel_path):
                messagebox.showerror("错误", "请先选择有效的Excel文件")
                return

            # 获取选中的文件
            selected_indices = self.file_listbox.curselection()
            if not selected_indices:
                messagebox.showwarning("警告", "请选择至少一个文件")
                return

            selected_files = [os.path.basename(self.found_files[i]) for i in selected_indices]
            name_col = self.name_col_var.get()
            start_row = self.start_row_var.get()

            try:
                # 打开现有工作簿
                wb = openpyxl.load_workbook(excel_path)

                # 获取选定的工作表
                sheet = self.get_worksheet(wb)

                # 填充文件名到指定单元格
                for i, filename in enumerate(selected_files):
                    sheet[f"{name_col}{start_row + i}"] = filename

                # 保存Excel文件
                wb.save(excel_path)
                self.log(f"文件名已填充到Excel: 工作表 '{sheet.title}' 从 {name_col}{start_row} 开始")
                messagebox.showinfo("成功", f"文件名已成功填充到工作表 '{sheet.title}' 的列 {name_col}")
            except Exception as e:
                messagebox.showerror("错误", f"填充文件名失败: {str(e)}")
                self.log(f"填充文件名失败: {str(e)}", logging.ERROR)

    def fill_texts(self):
        """精确匹配版文本填充 - 严格按文件名对应"""
        if not self.results:
            messagebox.showwarning("警告", "请先生成文本")
            return

        excel_path = self.excel_entry.get()
        if not excel_path or not os.path.exists(excel_path):
            messagebox.showerror("错误", "请先选择有效的Excel文件")
            return

        name_col = self.name_col_var.get().upper()  # 确保列名大写
        text_col = self.text_col_var.get().upper()
        start_row = int(self.start_row_var.get())  # 确保为整数

        try:
            wb = openpyxl.load_workbook(excel_path)
            sheet = self.get_worksheet(wb)

            # 创建精确文件名映射（保留原始大小写）
            result_map = {
                os.path.basename(result['file']): result['text']  # 使用完整文件名
                for result in self.results
            }

            matched_count = 0
            missing_files = []
            filled_rows = []

            # 遍历Excel行（严格按物理行号）
            for row_idx in range(start_row, sheet.max_row + 1):
                excel_filename = str(sheet[f"{name_col}{row_idx}"].value or "").strip()
                if not excel_filename:
                    continue

                # 精确匹配（包括扩展名）
                if excel_filename in result_map:
                    sheet[f"{text_col}{row_idx}"] = result_map[excel_filename]
                    matched_count += 1
                    filled_rows.append(row_idx)
                    result_map.pop(excel_filename)
                else:
                    missing_files.append(excel_filename)

            # 保存文件（带冲突处理）
            try:
                wb.save(excel_path)
            except PermissionError:
                backup_path = excel_path.replace(".xlsx", "_backup.xlsx")
                wb.save(backup_path)
                messagebox.showwarning("警告",
                                       f"原文件被占用，已保存到:\n{backup_path}")

            # 构建结果报告
            result_msg = [
                f"▸ 成功填充 {matched_count} 行",
                f"▸ 填充位置: 行 {min(filled_rows)}-{max(filled_rows)}" if filled_rows else "",
                f"▸ 未匹配文件: {len(missing_files)} 个" if missing_files else "",
                f"▸ 未使用结果: {len(result_map)} 条" if result_map else ""
            ]
            messagebox.showinfo("完成", "\n".join(filter(None, result_msg)))

        except Exception as e:
            messagebox.showerror("错误", f"填充失败:\n{str(e)}")
            self.log(f"填充错误 @ {excel_path}: {str(e)}", logging.ERROR)

    def compare_texts(self):
            """增强版文本比对功能 - 使用选定工作表"""
            excel_path = self.excel_entry.get()
            if not excel_path or not os.path.exists(excel_path):
                messagebox.showerror("错误", "请先选择有效的Excel文件")
                return

            text_col = self.text_col_var.get()
            compare_col = self.compare_col_var.get()
            similarity_threshold = self.similarity_var.get()
            start_row = self.start_row_var.get()

            try:
                # 打开Excel文件
                wb = openpyxl.load_workbook(excel_path)

                # 获取选定的工作表
                sheet = self.get_worksheet(wb)

                differences = 0
                diff_details = []

                # 遍历行进行文本比对
                row = start_row
                while sheet[f"{text_col}{row}"].value is not None:
                    text1 = str(sheet[f"{text_col}{row}"].value or "")
                    text2 = str(sheet[f"{compare_col}{row}"].value or "")

                    # 计算相似度
                    similarity = difflib.SequenceMatcher(None, text1, text2).ratio()

                    if similarity < similarity_threshold:
                        differences += 1
                        diff_details.append(f"行 {row}: 相似度 {similarity:.2f} < {similarity_threshold}")

                        # 标记差异行
                        sheet[f"{text_col}{row}"].font = Font(color="FF0000")  # 红色
                        sheet[f"{compare_col}{row}"].font = Font(color="FF0000")  # 红色

                    row += 1

                # 保存标记后的Excel
                wb.save(excel_path)

                # 显示结果
                self.log(f"工作表 '{sheet.title}' 文本比对完成，发现 {differences} 处差异")
                if differences > 0:
                    self.log("差异详情:")
                    for detail in diff_details:
                        self.log(f"  {detail}")

                messagebox.showinfo("完成",
                                    f"工作表 '{sheet.title}' 文本比对完成\n发现 {differences} 处显著差异 (阈值: {similarity_threshold})")
            except Exception as e:
                messagebox.showerror("错误", f"文本比对失败: {str(e)}")
                self.log(f"文本比对失败: {str(e)}", logging.ERROR)

    def export_report(self):
        """导出报告到日志文件（包含模型信息）"""
        if not self.results:
            messagebox.showwarning("警告", "没有可导出的结果")
            return

        # 准备日志文件名
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        first_file = os.path.splitext(os.path.basename(self.results[0]['file']))[0]
        model_name = self.model_var.get().replace("/", "_")  # 替换特殊字符
        log_name = f"{timestamp}_{model_name}_{first_file}_{len(self.results)}.txt"

        # 询问保存位置
        filetypes = [("日志文件", "*.txt"), ("所有文件", "*.*")]
        path = filedialog.asksaveasfilename(
            filetypes=filetypes,
            defaultextension=".txt",
            initialfile=log_name,
            initialdir=self.log_dir
        )

        if not path:
            return

        try:
            # 准备日志内容
            log_content = [
                f"语音识别报告 - {time.strftime('%Y-%m-%d %H:%M:%S')}",
                "=" * 60,
                f"模型名称: {self.model_var.get()}",
                f"模型类型: {self.models.get(self.model_var.get(), {}).get('engine', '未知')}",
                f"模型语言: {self.model_languages.get(self.model_var.get(), '未知')}",
                f"音频文件夹: {self.folder_entry.get()}",
                f"Excel文件: {self.excel_entry.get() or '未设置'}",
                f"处理文件数: {len(self.results)}",
                f"成功识别数: {len([r for r in self.results if r['text']])}",
                "\n详细识别结果:",
                "=" * 60
            ]

            # 添加每个文件的处理结果
            for idx, result in enumerate(self.results, 1):
                status = "成功" if result['text'] else "失败"
                log_content.append(
                    f"{idx}. 文件名: {result['file']}\n"
                    f"   状态: {status}\n"
                    f"   文本长度: {len(result['text'])}字符\n"
                    f"   识别结果:\n{result['text']}\n"
                    f"{'-' * 40}"
                )

            # 写入文件
            with open(path, 'w', encoding='utf-8') as f:
                f.write("\n".join(log_content))

            self.log(f"报告已导出: {path}")
            messagebox.showinfo("导出成功", f"报告已保存到:\n{path}")

        except Exception as e:
            messagebox.showerror("导出失败", f"导出报告时出错: {str(e)}")
            self.log(f"导出失败: {str(e)}", logging.ERROR)

    def clear_log(self):
            """清空日志"""
            self.log_text.config(state=tk.NORMAL)
            self.log_text.delete(1.0, tk.END)
            self.log_text.config(state=tk.DISABLED)
            self.log("日志已清空")


if __name__ == "__main__":
    # 配置根日志器
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler("audio_to_text.log", encoding="utf-8"),
            logging.StreamHandler()
        ]
    )

    try:
        root = tk.Tk()
        app = AudioToTextTool(root)


        def on_closing():
            if hasattr(app, 'is_processing') and app.is_processing:
                if messagebox.askokcancel("退出", "处理正在进行中，确定要退出吗?"):
                    app.is_processing = False
                    root.destroy()
            else:
                root.destroy()


        root.protocol("WM_DELETE_WINDOW", on_closing)
        root.mainloop()

    except Exception as e:
        logging.error(f"程序启动失败: {str(e)}", exc_info=True)
        messagebox.showerror("致命错误", f"程序启动失败:\n{str(e)}")